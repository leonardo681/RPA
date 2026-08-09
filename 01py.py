import os
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple

try:
    import requests
except ImportError as exc:
    raise SystemExit(f"Biblioteca ausente: {exc}. Instale com: pip install requests")

try:
    import pyautogui
except ImportError as exc:
    raise SystemExit(f"Biblioteca ausente: {exc}. Instale com: pip install pyautogui")

try:
    from openpyxl import Workbook
except ImportError as exc:
    raise SystemExit(f"Biblioteca ausente: {exc}. Instale com: pip install openpyxl")


WORKSPACE_ROOT = Path(__file__).resolve().parent
RELATORIOS_DIR = WORKSPACE_ROOT / "relatorios"

pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.8


def criar_planilha(nome_arquivo: str = "dados_site.xlsx", nome_planilha: str = "Dados") -> Tuple[Workbook, object, str]:
    """Cria uma nova planilha Excel dentro da pasta relatorios do projeto."""
    RELATORIOS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = nome_planilha
    caminho = RELATORIOS_DIR / nome_arquivo
    return wb, ws, str(caminho)


def salvar_planilha(wb: Workbook, caminho: str) -> str:
    """Salva a planilha Excel."""
    caminho_obj = Path(caminho)
    caminho_obj.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_obj)
    return str(caminho_obj)


def coletar_valores_site() -> List[Dict[str, str]]:
    """Acessa um site e coleta três valores distintos para gravar na planilha."""
    url = "https://economia.awesomeapi.com.br/json/last/USD-BRL,EUR-BRL,BTC-BRL"
    response = requests.get(url, timeout=20)
    response.raise_for_status()
    dados = response.json()

    registros: List[Dict[str, str]] = []
    for chave, nome in [("USDBRL", "Dólar"), ("EURBRL", "Euro"), ("BTCBRL", "Bitcoin")]:
        item = dados.get(chave, {})
        registros.append(
            {
                "nome": nome,
                "valor_compra": item.get("bid", "N/A"),
                "valor_venda": item.get("ask", "N/A"),
                "maxima": item.get("high", "N/A"),
                "data_hora": item.get("create_date", datetime.now().strftime("%d/%m/%Y %H:%M")),
            }
        )

    return registros


def escrever_dados_planilha(ws, registros: List[Dict[str, str]]) -> None:
    """Escreve cabeçalhos e os valores em colunas separadas na planilha."""
    cabecalhos = ["Nome", "Valor Compra", "Valor Venda", "Máxima", "Data/Hora"]
    for coluna, titulo in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=coluna, value=titulo)

    for linha, registro in enumerate(registros, start=2):
        ws.cell(row=linha, column=1, value=registro["nome"])
        ws.cell(row=linha, column=2, value=registro["valor_compra"])
        ws.cell(row=linha, column=3, value=registro["valor_venda"])
        ws.cell(row=linha, column=4, value=registro["maxima"])
        ws.cell(row=linha, column=5, value=registro["data_hora"])


def abrir_excel(caminho_arquivo: str) -> None:
    """Abre o arquivo Excel na interface do Windows para visualização."""
    print("Abrindo o Excel...")
    os.startfile(caminho_arquivo)
    time.sleep(5)


def executar_fluxo() -> str:
    """Executa todo o fluxo: acessa o site, coleta os valores e grava na planilha Excel."""
    wb, ws, caminho_arquivo = criar_planilha()
    registros = coletar_valores_site()
    escrever_dados_planilha(ws, registros)

    caminho_salvo = salvar_planilha(wb, caminho_arquivo)
    print(f"Planilha criada em: {caminho_salvo}")
    print("Valores capturados:")
    for registro in registros:
        print(registro)

    abrir_excel(caminho_salvo)
    return caminho_salvo


if __name__ == "__main__":
    executar_fluxo()
