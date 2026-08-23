import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import selenium 
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys
import pyautogui as p 
import datetime as dt
import rpa as r
import pandas as pd
import time
import os
import io
from openpyxl.drawing.image import Image as OpenpyxlImage
from dotenv import load_dotenv
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup

# Carrega as variáveis do arquivo .env para o ambiente Python
load_dotenv()

# Captura os valores do .env
URL_SITE = os.getenv("URL_SITE")
EMAIL_USUARIO = os.getenv("EMAIL_USUARIO")
SENHA_USUARIO = os.getenv("SENHA_USUARIO")

def criar_pasta_relatorios(pasta='relatorios'):
    """Garante que a pasta para salvar os arquivos exista."""
    if not os.path.exists(pasta):
        os.makedirs(pasta)

def wait_for_element(element_selector, timeout=15):
    """
    Aguarda de forma reativa até que um elemento apareça na página.
    Retorna True se o elemento for encontrado, ou gera uma exceção no tempo limite.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        if r.present(element_selector):
            return True
        time.sleep(0.2)  # Verifica a cada 200 milissegundos
    raise Exception(f"Timeout: O elemento '{element_selector}' não carregou dentro de {timeout} segundos.")

def realizar_login(email, senha, url=URL_SITE):
    """
    Função responsável por inicializar o navegador, acessar o site e
    executar o fluxo completo de autenticação via RPA.
    """
    # 1. Inicializa o ambiente RPA
    r.init(visual_automation=True, chrome_browser=True)
    r.url(url)
    
    # 2. Maximiza a janela ativa
    janela = p.getActiveWindow()
    if janela:
        janela.maximize()

    # 3. Fluxo de Login usando os seletores diretos
    print("Iniciando processo de login...")
    
    # Clica em 'Entrar'
    r.click('Entrar')

    # Preenche o e-mail e avança
    r.type('#identifier-field', email)
    r.click('Continuar')

    # Preenche a senha e avança
    r.type('#password-field', senha)
    r.click('Continuar')
    
    print("Login concluído com sucesso!")
    
def teste(url='https://www.tabelatacoonline.com.br/'):
    """
    Função responsável por inicializar o navegador, acessar o site e
    executar o fluxo completo de autenticação via RPA.
    """
    # 1. Inicializa o ambiente RPA
    r.init(visual_automation=True, chrome_browser=True)
    r.url(url)
    
    # 2. Maximiza a janela ativa
    janela = p.getActiveWindow()
    if janela:
        janela.maximize()

    # 3. Fluxo de Login usando os seletores diretos
    print("pagina acessada com sucesso!")
    
def rolar_pagina(modo='pixels', pixels=700):
   
    if modo == 'fim':
        time.sleep(2)  # Aguarda a animação da rolagem terminar
        r.dom("window.scrollTo({top: document.body.scrollHeight, behavior: 'smooth'});")
        time.sleep(3)  # Aguarda a animação da rolagem terminar
        r.dom("window.scrollTo({top: 0, behavior: 'smooth'});")
        time.sleep(3)  # Aguarda a animação da rolagem terminar
        rolar_pagina(modo='pixels', pixels=700)

    elif modo == 'pixels':
        # Rola uma quantidade exata de pixels para baixo
        time.sleep(2)  # Aguarda a animação da rolagem terminar
        r.dom(f"window.scrollBy({{top: {pixels}, behavior: 'smooth'}});")

def acessar_categoria(nome_categoria):
    """
    Clica na categoria desejada e aguarda o carregamento da nova página.
    """
    print(f"Clicando na categoria: {nome_categoria}...")
    r.click(nome_categoria)
    print("Página carregada com sucesso!")
        
def extrair_alimentos_apenas_python(nome_arquivo_csv=os.path.join('relatorios', 'alimentos_preparados.csv')):
    criar_pasta_relatorios()
    print("Rolando a página para carregar todos os dados...")
    rolar_pagina(modo='fim')

    print("Capturando HTML completo...")
    html_pagina = r.read('page')
    
    soup = BeautifulSoup(html_pagina, 'html.parser')
    cards = soup.find_all('a', href=lambda href: href and '/tabela-nutricional/' in href)
    
    dados_alimentos = []

    for card in cards:
        texto_card = card.text.strip()
        
        if 'Conta PRO' in texto_card:
            continue
            
        tag_nome = card.find('p', title=True)
        nome = tag_nome['title'] if tag_nome else ''
        
        paragrafos = [p.text.strip() for p in card.find_all('p')]
        
        energia = ''
        carboidratos = ''
        proteinas = ''
        
        for idx, texto in enumerate(paragrafos):
            if 'Energia:' in texto and idx + 1 < len(paragrafos):
                energia = paragrafos[idx + 1]
            elif 'Carboidratos:' in texto and idx + 1 < len(paragrafos):
                carboidratos = paragrafos[idx + 1]
            elif 'Proteínas:' in texto and idx + 1 < len(paragrafos):
                proteinas = paragrafos[idx + 1]

        if nome and energia and 'kcal' in energia:
            link_href = card.get('href', '')
            link_completo = f"https://www.tabelatacoonline.com.br{link_href}" if link_href.startswith('/') else link_href
            
            dados_alimentos.append({
                'Alimento': nome,
                'Energia': energia,
                'Carboidratos': carboidratos,
                'Proteínas': proteinas,
                'Link': link_completo
            })

    print(f"✓ Sucesso! {len(dados_alimentos)} alimentos gratuitos extraídos com dados válidos.")

    if dados_alimentos:
        df = pd.DataFrame(dados_alimentos)
        df.to_csv(nome_arquivo_csv, index=False, encoding='utf-8-sig')
        print(f"✓ Arquivo '{nome_arquivo_csv}' gerado com sucesso!")
    else:
        print("Nenhum dado válido encontrado para salvar.")

    return dados_alimentos

def salvar_em_excel(
    nome_csv=os.path.join('relatorios', 'alimentos_preparados.csv'),
    nome_excel=os.path.join('relatorios', 'alimentos_preparados.xlsx')
):
    """
    Lê o CSV e converte para Excel dentro da pasta 'relatorios' com rankings e gráficos.
    """
    criar_pasta_relatorios()
    if not os.path.exists(nome_csv):
        print(f"Erro: O arquivo {nome_csv} não foi encontrado para converter em Excel.")
        return

    print("Convertendo CSV para Excel (.xlsx) e gerando gráficos...")
    
    df = pd.read_csv(nome_csv)
    
    for col in ['Energia', 'Carboidratos', 'Proteínas']:
        if col in df.columns:
            df[f'{col}_num'] = (
                df[col].astype(str)
                .str.replace(r'[^\d,.]', '', regex=True)
                .str.replace(',', '.')
            )
            df[f'{col}_num'] = pd.to_numeric(df[f'{col}_num'], errors='coerce').fillna(0)

    df_geral = df.drop(columns=['Energia_num', 'Carboidratos_num', 'Proteínas_num'], errors='ignore')

    def preparar_ranking(coluna_sort):
        return (
            df.sort_values(by=coluna_sort, ascending=False)
            .drop(columns=['Link', 'Energia_num', 'Carboidratos_num', 'Proteínas_num'], errors='ignore')
        )

    df_energia = preparar_ranking('Energia_num')
    df_carbo = preparar_ranking('Carboidratos_num')
    df_proteina = preparar_ranking('Proteínas_num')

    abas_config = {
        'Geral': {'df': df_geral, 'grafico_col': None, 'unidade': ''},
        'Ranking Energia': {'df': df_energia, 'grafico_col': 'Energia_num', 'unidade': 'kcal', 'cor': '#e67e22'},
        'Ranking Carboidratos': {'df': df_carbo, 'grafico_col': 'Carboidratos_num', 'unidade': 'g', 'cor': '#2ecc71'},
        'Ranking Proteínas': {'df': df_proteina, 'grafico_col': 'Proteínas_num', 'unidade': 'g', 'cor': '#3498db'}
    }

    wb = Workbook()
    
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True)

    for index, (nome_aba, config) in enumerate(abas_config.items()):
        if index == 0:
            ws = wb.active
            ws.title = nome_aba
        else:
            ws = wb.create_sheet(title=nome_aba)

        df_aba = config['df']
        colunas = list(df_aba.columns)
        ws.append(colunas)

        for col_num in range(1, len(colunas) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for linha in df_aba.itertuples(index=False):
            ws.append(list(linha))

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        col_num_ref = config['grafico_col']
        if col_num_ref:
            top10 = df.sort_values(by=col_num_ref, ascending=False).head(10)
            
            plt.figure(figsize=(8, 4.5))
            plt.barh(top10['Alimento'][::-1], top10[col_num_ref][::-1], color=config['cor'])
            plt.title(f'Top 10 - {nome_aba.replace("Ranking ", "")}', fontsize=12, fontweight='bold')
            plt.xlabel(f'Quantidade ({config["unidade"]})')
            plt.tight_layout()

            img_buf = io.BytesIO()
            plt.savefig(img_buf, format='png', dpi=120)
            img_buf.seek(0)
            plt.close()

            img = OpenpyxlImage(img_buf)
            ws.add_image(img, 'F2')

    wb.save(nome_excel)
    print(f"✓ Arquivo Excel '{nome_excel}' criado na pasta relatórios!")
    
def abrir_excel(nome_excel=os.path.join('relatorios', 'alimentos_preparados.xlsx')):
    """Abre o arquivo Excel localizado na pasta relatorios."""
    if os.path.exists(nome_excel):
        print(f"Abrindo o arquivo '{nome_excel}'...")
        os.startfile(os.path.abspath(nome_excel))
    else:
        print(f"Erro: O arquivo '{nome_excel}' não foi encontrado para ser aberto.")

def fechar_navegador():
    """
    Encerra a sessão do RPA e fecha a janela do navegador aberta pela automação.
    """
    print("Fechando o navegador...")
    r.close()
    print("✓ Navegador fechado com sucesso!")

# --- EXECUÇÃO DO MÉTODO ---
if __name__ == "__main__":
    
    realizar_login(EMAIL_USUARIO, SENHA_USUARIO)
    rolar_pagina(modo='pixels', pixels=700)
    acessar_categoria('//a[.//p[text()="Alimentos preparados"]]')
    alimentos = extrair_alimentos_apenas_python('alimentos_preparados.csv')
    salvar_em_excel('alimentos_preparados.csv', 'alimentos_preparados.xlsx')
    abrir_excel('alimentos_preparados.xlsx')
    fechar_navegador()

    