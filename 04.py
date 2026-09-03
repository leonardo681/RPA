import openpyxl 
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage

import selenium 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import Chrome
from selenium.webdriver.common.keys import Keys

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


import os
import time
import urllib.parse
import rpa as r
import pyautogui as p
import pyautogui as p 
import urllib.parse
import datetime as dt
import rpa as r
import pandas as pd
import time
import os
import io
from dotenv import load_dotenv
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
from pixqrcodegen import Payload

# Carrega as variáveis do arquivo .env para o ambiente Python
load_dotenv()

# Captura os valores do .env
URL_SITE = os.getenv("URL_SITE")
EMAIL_USUARIO = os.getenv("EMAIL_USUARIO")
SENHA_USUARIO = os.getenv("SENHA_USUARIO")

def wait_for_element(element_selector, timeout=15):
    """
    Aguarda de forma reativa até que um elemento apareça na página.
    Retorna True se o elemento for encontrado, ou gera uma exceção no tempo limite.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        if r.present(element_selector):
            return True
        time.sleep(0.2)  # Verifica a cada 200 milissegundos
    raise Exception(f"Timeout: O elemento '{element_selector}' não carregou dentro de {timeout} segundos.")

def abrir_navegador():
    """
    Inicializa o navegador Chrome com a biblioteca RPA.
    """
    r.init(visual_automation=True, chrome_browser=True)
    janela = p.getActiveWindow()
    if janela:
        janela.maximize()

def abrir_whatsapp_web():
    """
    Navega até o site do WhatsApp Web na sessão do navegador.
    """
    print("Acessando o WhatsApp Web...")
    r.url('https://web.whatsapp.com')
    print("✓ WhatsApp Web carregado!")

def gerar_planilha_clientes(
    nome_excel=os.path.join('relatorios', 'clientes_cobranca.xlsx')
):
    """Gera uma planilha Excel com dados fictícios de cobrança de clientes."""
    print("Gerando planilha de clientes...")

    # Garante que a pasta destino existe
    if not os.path.exists(os.path.dirname(nome_excel)):
        os.makedirs(os.path.dirname(nome_excel))

    # Dados fictícios
    dados = [
        {
            'nome': 'Carlos Eduardo Silva',
            'telefone': '5511987654321',
            'email': 'carlos.silva@email.com',
            'data de vencimento': '15/10/2026',
            'valor a ser pago': 'R$ 250,50',
        },
        {
            'nome': 'Mariana Souza',
            'telefone': '5521998765432',
            'email': 'mariana.souza@email.com',
            'data de vencimento': '20/10/2026',
            'valor a ser pago': 'R$ 480,00',
        },
        {
            'nome': 'Roberto Alves',
            'telefone': '5531976543210',
            'email': 'roberto.alves@email.com',
            'data de vencimento': '25/10/2026',
            'valor a ser pago': 'R$ 135,90',
        },
    ]

    # Criando o DataFrame e salvando em Excel
    df = pd.DataFrame(dados)
    df.to_excel(nome_excel, index=False)

    print(f"✓ Planilha gerada com sucesso em '{nome_excel}'!")
    return nome_excel

def coletar_dados_cobranca(nome_arquivo='clientes_cobranca.xlsx', pasta='relatorios'):
    """
    Lê a planilha de clientes na pasta relatorios e retorna um dicionário 
    contendo TODOS os campos cadastrados de cada cliente, chaveados pelo telefone.
    """
    caminho_excel = os.path.join(pasta, nome_arquivo)

    if not os.path.exists(caminho_excel):
        print(f"⚠️ Atenção: A planilha '{caminho_excel}' não foi encontrada.")
        return {}

    print(f"Lendo todos os dados da planilha: '{caminho_excel}'...")
    
    # Lê a planilha
    df = pd.read_excel(caminho_excel)

    dados_cobranca = {}

    for _, linha in df.iterrows():
        # Define o telefone como chave do dicionário
        telefone = str(linha['telefone']).replace('.0', '').strip()
        
        # Mapeia dinamicamente todas as colunas presentes na linha
        info_cliente = {}
        for coluna in df.columns:
            valor = linha[coluna]
            info_cliente[coluna] = str(valor).strip() if pd.notna(valor) else ""

        dados_cobranca[telefone] = info_cliente

    print(f"✓ Sucesso! {len(dados_cobranca)} cliente(s) carregado(s) com todos os dados.")
    return dados_cobranca

def gerar_boleto_cliente(info_cliente, pasta_destino='relatorios'):
    """
    Gera um PDF personalizado de cobrança/boleto para um cliente específico.
    """
    if not os.path.exists(pasta_destino):
        os.makedirs(pasta_destino)

    nome = info_cliente.get('nome', 'Cliente')
    telefone = info_cliente.get('telefone', '')
    email = info_cliente.get('email', '')
    vencimento = info_cliente.get('data de vencimento', 'A vencer')
    valor = info_cliente.get('valor a ser pago', 'R$ 0,00')

    nome_pdf = f"boleto_{telefone}.pdf"
    caminho_pdf = os.path.join(pasta_destino, nome_pdf)

    doc = SimpleDocTemplate(
        caminho_pdf,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    story = []
    styles = getSampleStyleSheet()

    # Estilos do documento
    style_header = ParagraphStyle('Header', fontName='Helvetica-Bold', fontSize=20, textColor=colors.HexColor('#1E3A8A'), spaceAfter=5)
    style_sub = ParagraphStyle('Sub', fontName='Helvetica', fontSize=12, textColor=colors.HexColor('#555555'), spaceAfter=15)
    style_title = ParagraphStyle('Title', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor('#333333'), spaceAfter=4)
    style_text = ParagraphStyle('Text', fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#444444'))
    style_bold = ParagraphStyle('Bold', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor('#111111'))

    # Cabeçalho
    story.append(Paragraph("DEMONSTRATIVO DE COBRANÇA", style_header))
    story.append(Paragraph("Aviso de Vencimento de Fatura", style_sub))
    story.append(Spacer(1, 10))

    # Dados do Cliente e Cobrança
    dados_tabela = [
        [Paragraph("DADOS DO CLIENTE", style_title), Paragraph("DETALHES DO PAGAMENTO", style_title)],
        [Paragraph(f"<b>Nome:</b> {nome}<br/><b>E-mail:</b> {email}<br/><b>Tel:</b> {telefone}", style_text),
         Paragraph(f"<b>Vencimento:</b> {vencimento}<br/><b>Valor Total:</b> {valor}", style_text)]
    ]
    tabela = Table(dados_tabela, colWidths=[250, 250])
    tabela.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(tabela)
    story.append(Spacer(1, 20))

    # Instruções
    story.append(Paragraph("<b>Instruções de Pagamento:</b>", style_bold))
    story.append(Paragraph("Utilize a chave Pix registrada no nosso sistema para efetuar o pagamento até a data de vencimento.", style_text))
    story.append(Spacer(1, 15))

    doc.build(story)
    return os.path.abspath(caminho_pdf)

def enviar_boletos_whatsapp(carteira_clientes, pasta_destino='relatorios'):
    """
    Gera o PDF individual de cada cliente e envia via WhatsApp Web.
    Garante o foco no chat para enviar a mensagem e utiliza seletores atualizados para o anexo.
    """
    if not carteira_clientes:
        print("Nenhum cliente fornecido para envio.")
        return

    print("Iniciando processo de geração e envio de boletos...")

    for telefone, info in carteira_clientes.items():
        nome = info.get('nome', 'Cliente')
        valor = info.get('valor a ser pago', '')
        vencimento = info.get('data de vencimento', '')

        # 1. Gera o PDF do cliente
        caminho_pdf = gerar_boleto_cliente(info, pasta_destino)
        print(f"\n[+] PDF gerado para {nome}: {caminho_pdf}")

        # 2. Formata e codifica a mensagem para a URL
        mensagem = f"Olá {nome}, tudo bem? Segue em anexo o seu demonstrativo no valor de {valor} com vencimento para {vencimento}."
        mensagem_codificada = urllib.parse.quote(mensagem)
        
        # 3. Abre a conversa direta no WhatsApp Web
        url_wa = f"https://web.whatsapp.com/send?phone={telefone}&text={mensagem_codificada}"
        r.url(url_wa)
        
        print(f"Aguardando carregamento da conversa com {nome} ({telefone})...")
        time.sleep(10)  # Tempo para o WhatsApp carregar a conversa completamente

        # 4. Validação: Verifica se o número existe no WhatsApp
        numero_invalido = r.present('//div[contains(text(), "inválido")]') or \
                          r.present('//div[contains(text(), "invalid")]') or \
                          r.present('//div[contains(text(), "não está no WhatsApp")]')

        if numero_invalido:
            print(f"⚠️ O número {telefone} do cliente '{nome}' não existe no WhatsApp. Pulo efetuado.")
            if r.present('//button[contains(., "OK")]'):
                r.click('//button[contains(., "OK")]')
            continue

        print(f"✓ Número {telefone} confirmado!")

        # 5. Envia a MENSAGEM DE TEXTO (Clica na caixa de texto para dar foco antes de enviar)
        seletor_caixa_texto = '//footer//div[@contenteditable="true"]'
        seletor_botao_enviar_texto = '//span[@data-icon="send"] | //button[@aria-label="Enviar"] | //button[@title="Enviar"]'

        if r.present(seletor_caixa_texto):
            r.click(seletor_caixa_texto)
            time.sleep(1)
            
            if r.present(seletor_botao_enviar_texto):
                r.click(seletor_botao_enviar_texto)
            else:
                r.keyboard('[enter]')
            
            print(f"✓ Mensagem de texto enviada para {nome}!")
            time.sleep(3)
        else:
            print(f"⚠️ Caixa de texto não encontrada para {nome}.")

        # 6. Anexar e enviar o documento PDF
        # Seletores combinados para encontrar o botão de anexo (+)
        seletor_anexo = '//button[@title="Anexar"] | //button[@aria-label="Anexar"] | //span[@data-icon="plus"] | //span[@data-icon="attach-menu-plus"] | //div[@title="Anexar"]'
        
        try:
            if r.present(seletor_anexo):
                r.click(seletor_anexo)
                time.sleep(2)

                # Clica na opção 'Documento' do menu que abre
                seletor_doc = '//span[contains(text(), "Documento")] | //button[contains(., "Documento")] | //li[contains(., "Documento")]'
                if r.present(seletor_doc):
                    r.click(seletor_doc)
                    time.sleep(2)

                    # Interage com a janela de arquivos do Windows
                    p.write(caminho_pdf)
                    p.press('enter')
                    time.sleep(3)

                    # Clica no botão de enviar o PDF na tela de pré-visualização
                    seletor_enviar_doc = '//span[@data-icon="send"] | //div[@aria-label="Enviar"] | //button[@aria-label="Enviar"]'
                    if r.present(seletor_enviar_doc):
                        r.click(seletor_enviar_doc)
                    else:
                        r.keyboard('[enter]')
                    
                    print(f"✓ Boleto PDF enviado com sucesso para {nome}!")
                    time.sleep(3)
                else:
                    print(f"⚠️ Opção 'Documento' não foi encontrada no menu de anexos.")
            else:
                print(f"⚠️ Não foi possível localizar o botão de anexo para {nome}.")

        except Exception as e:
            print(f"Erro ao anexar arquivo para {nome}: {e}")

    print("\n✓ Processo de envio concluído!")



# --- EXECUÇÃO DO MÉTODO ---
if __name__ == "__main__":
    
    abrir_navegador()
    abrir_whatsapp_web()
    #gerar_planilha_clientes()
    carteira = coletar_dados_cobranca()
    if carteira:
        enviar_boletos_whatsapp(carteira)
